
import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from datetime import datetime

def limpiar_dominio(valor):
    if not isinstance(valor, str):
        return valor
    return ''.join(filter(str.isalnum, valor.upper()))

def normalizar(valor):
    if valor is None:
        return ""
    return str(valor).strip().lower()

def validar_fecha(valor):
    if pd.isna(valor) or str(valor).strip() == "":
        return "", True, ""
    try:
        if isinstance(valor, datetime):
            return valor.strftime("%d/%m/%Y"), True, ""
        elif isinstance(valor, str):
            val = valor.lower().strip()
            meses = {
                "enero": "01", "febrero": "02", "marzo": "03", "abril": "04",
                "mayo": "05", "junio": "06", "julio": "07", "agosto": "08",
                "septiembre": "09", "setiembre": "09", "octubre": "10",
                "noviembre": "11", "diciembre": "12"
            }
            for mes_texto, mes_num in meses.items():
                if mes_texto in val:
                    val = val.replace(mes_texto, mes_num)
            val = val.replace("-", "/").replace(".", "/")
            fecha = pd.to_datetime(val, dayfirst=True, errors='coerce')
            if pd.isna(fecha):
                raise ValueError("No se pudo convertir")
            return fecha.strftime("%d/%m/%Y"), True, ""
        fecha = pd.to_datetime(valor, dayfirst=True, errors='raise')
        return fecha.strftime("%d/%m/%Y"), True, ""
    except Exception as e:
        return valor, False, f"Fecha inválida: {e}"

st.title("Validador Excel Vehículos")

file = st.file_uploader("Subí el archivo Excel", type=["xlsx"])

if file:
    wb = load_workbook(file)
    ws = wb.active

    encabezados = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    encabezados_norm = [normalizar(e) for e in encabezados]

    errores = []
    corregidos = []
    cambios_por_columna = {}
    valores_unicos = {"dominio": set()}
    fill_red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    font_white = Font(color="FFFFFF", bold=True)

    columnas_fecha = [
        "vto póliza", "vto cédula", "vto vtv", "vto ruta",
        "vto gnc", "vto cilindro gnc", "vto senasa"
    ]

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            col_idx = cell.column - 1
            col_name = encabezados_norm[col_idx]
            val = cell.value
            col_title = encabezados[col_idx]

            if col_name == "dominio":
                nuevo = limpiar_dominio(val)
                if nuevo in valores_unicos["dominio"]:
                    cell.fill = fill_red
                    cell.font = font_white
                    errores.append((cell.row, col_title, val, "Duplicado"))
                else:
                    valores_unicos["dominio"].add(nuevo)
                    if nuevo != val:
                        corregidos.append((cell.row, col_title, val, nuevo))
                        cambios_por_columna[col_title] = cambios_por_columna.get(col_title, 0) + 1
                        cell.value = nuevo
                continue

            
            if col_name == "dominio":
                nuevo = limpiar_dominio(val)
                if nuevo in valores_unicos["dominio"]:
                    cell.fill = fill_red
                    cell.font = font_white
                    errores.append((cell.row, col_title, val, "Duplicado"))
                else:
                    valores_unicos["dominio"].add(nuevo)
                    if nuevo != val:
                        corregidos.append((cell.row, col_title, val, nuevo))
                        cambios_por_columna[col_title] = cambios_por_columna.get(col_title, 0) + 1
                        cell.value = nuevo
                continue

            if col_name in columnas_fecha:
                nuevo, ok, msg = validar_fecha(val)
                if not ok:
                    cell.fill = fill_red
                    cell.font = font_white
                    errores.append((cell.row, col_title, val, msg))
                elif nuevo != val:
                    corregidos.append((cell.row, col_title, val, nuevo))
                    cambios_por_columna[col_title] = cambios_por_columna.get(col_title, 0) + 1
                    cell.value = nuevo

    st.markdown(f"### 📊 Resumen general")
    st.info(f"Se detectaron **{len(errores)} errores** y se corrigieron automáticamente **{len(corregidos)} valores**.")

    with st.expander("📋 Ver detalles de cambios"):
        if errores:
            st.markdown("### ❌ No corregidos")
            st.dataframe(pd.DataFrame(errores, columns=["Fila", "Columna", "Valor original", "Observación"]))
        if corregidos:
            st.markdown("### ✅ Corregidos")
            st.dataframe(pd.DataFrame(corregidos, columns=["Fila", "Columna", "Valor original", "Valor corregido"]))

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    st.download_button(
        "📥 Descargar Excel Validado",
        data=output,
        file_name="vehiculos_validado.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
